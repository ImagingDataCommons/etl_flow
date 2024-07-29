from google.cloud import bigquery
import json
import re
import pandas as pd
import numpy as np
import re
import sys
from os import path, listdir,mkdir
import zipfile
import clinical.acrin_forms
import shutil
from pathlib import Path
import hashlib
import pytz
from datetime import datetime
import os
from clinical.utils import getHist, read_clin_file, parseIspyDic
from docx2python import docx2python
import openpyxl
#import copy.deepcopy
from python_settings import settings
import settings as etl_settings
settings.configure(etl_settings)
assert settings.configured


ORIGINAL_SRCS_PATH='./clinical/downloads/downloads_'+str(settings.CURRENT_VERSION)+'/'
#ORIGINAL_SRCS_PATH= '/Users/george/fed/actcianable/output/clinical_files/'
#NOTES_PATH = '/Users/george/fed/actcianable/output/'
NOTES_PATH = './clinical/'
DEFAULT_SUFFIX='clinical'
DEFAULT_DESCRIPTION='clinical data'


DEFAULT_DATASET = settings.BQ_CLIN_DATASET
DEFAULT_PROJECT = settings.DEV_PROJECT
CURRENT_VERSION = 'idc_v'+str(settings.CURRENT_VERSION)
LAST_VERSION = 'idc_v'+str(settings.PREVIOUS_VERSION)
LAST_DATASET = 'idc_v'+str(settings.PREVIOUS_VERSION)+'_clinical'
DESTINATION_FOLDER='./clinical/json/clin_idc_v'+str(settings.CURRENT_VERSION)+'/'
SOURCE_BATCH_COL='source_batch'
SOURCE_BATCH_LABEL='idc_provenance_source_batch'
DICOM_COL= 'dicom_patient_id'
DICOM_LABEL='idc_provenance_dicom_patient_id'
DATASET_PATH='bigquery-public-data.'+DEFAULT_DATASET
ARCHIVE_FOLDER = './clinical/archive/'

def get_md5(filenm):
  with open(filenm, 'rb') as file_to_check:
    # read contents of the file
    data = file_to_check.read()
    # pipe contents of the file through
    return(hashlib.md5(data).hexdigest())

def write_dataframe_to_json(path,nm,df):
  #headers = clinJson[coll]['headers']
  filenm=path+nm+'.json'
  f = open(filenm, 'w')
  cols=list(df.columns)
  nArr = []
  for i in range(len(cols)):
    col=df.columns[i]
    dtype=df.dtypes[i].name
    ntype=''
    if dtype=='object':
      ntype='str'
    elif dtype=='float64':
      ntype='float'
    elif dtype=='int64':
      ntype='int'
    elif dtype == 'datetime64[ns]':
      ntype='datetime'
      df[df.columns[i]] = df[df.columns[i]].astype(str)
    elif dtype == 'bool':
      ntype = 'bool'
    else:
      ntype = 'str'
      df[df.columns[i]] = df[df.columns[i]].astype(str)
    nArr.append([col, ntype])
  data = [{**row.dropna().to_dict()} for index, row in df.iterrows()]
  out={'schema':nArr, 'data':data}
  try:
    json.dump(out,f)
  except:
    pass
  f.close()





def write_clin_file(filenm, data):

  for curKey in clinJson:
    ndata = clinJson[curKey]
    f = open(filenm, 'w')
    json.dump(ndata, f)
    f.close()

def recastDataFrameTypes(df, ptId):
  for i in range(len(df.columns)):
    if not (i == ptId) and (df.dtypes[i].name == 'float64'):
      try:
        df[df.columns[i]] = df[df.columns[i]].astype('Int64')
      except:
        pass
    # make all not na objects strings
    if (df.dtypes[i].name == 'object'):
      try:
        df[df.columns[i]] = df[df.columns[i]].map(lambda a: a if pd.isna(a) else str(a))
      except:
        pass

def analyzeDataFrame(cdic):
  df = cdic['df']
  for i in range(len(df.columns)):
    try:
      uVals = list(df[df.columns[i]].unique())
      uVals = ["null" if pd.isna(x) else x for x in uVals]
    except:
      pass
    try:
      uVals.sort()
    except:
      pass
    try:
      if len(cdic['headers'][df.columns[i]])>0:
        cdic['headers'][df.columns[i]][0]['uniques']=uVals
        if (df.dtypes[i].name == 'float64') or (df.dtypes[i].name == 'Int64'):
          if (len(uVals)>0):
            cdic['headers'][df.columns[i]][0]['rng']=[float(uVals[0]),float(uVals[len(uVals)-1])]
            iii=1
    except:
      iii=1


def processSrc(fpath, colName, srcInfo, coll):
  attrs=[]
  filenm = fpath+colName+'/'+srcInfo['filenm']
  sheetNo = (0 if not 'sheet' in srcInfo else srcInfo['sheet'])
  patientIdRow = (0 if not ('patientIdRow') in srcInfo else srcInfo['patientIdRow'])
  rows = ([0] if not 'headrows' in srcInfo else srcInfo['headrows'])
  skipRows = (None if not 'skipRows' in srcInfo else srcInfo['skipRows'])
  skipCols = (None if not 'skipCols' in srcInfo else srcInfo['skipCols'])
  pivot= (False if not 'pivot' in srcInfo else srcInfo['pivot'])
  maxRow = (-1 if not 'maxRow' in srcInfo else srcInfo['maxRow'])
  filterRows = (None if not 'filterRows' in srcInfo else srcInfo['filterRows'])
  extension = path.splitext(filenm)[1]
  engine='xlrd'
  if extension == '.xlsx':
    engine= 'openpyxl'
  elif extension == '.xlsb':
    engine = 'pyxlsb'
  df=[]
  if extension == '.csv':
    df = pd.read_csv(filenm, keep_default_na=False)
    #df = df.head(100)
    sheetnm=''
  else:
    dfi = pd.read_excel(filenm, engine=engine, sheet_name=None, keep_default_na=False)
    sheetnm = list(dfi.keys())[sheetNo]
    df = dfi[sheetnm]
  if pivot:
    df = df.T
    rows =[rows[i]+1 for i in range(len(rows))]
    colList=list(df.columns)
    df.insert(0,'tmp',list(df.index))

  if skipCols is not None:
    df.drop(columns=[df.columns[i] for i in skipCols],inplace=True)

  for i in range(len(df.columns)):
    attrs.append([])

  for i in range(len(rows)):
    colVal=''
    ind = rows[i]
    if ind == 0:
      values = df.columns
    else:
      values=df.values[ind-1]
    for j in range(len(values)):
      val=values[j]
      if (i == len(rows)-1) or (not (str(val) == 'nan') and not ('Unnamed:' in str(val))):
        if 'headerformatspec' in srcInfo:
          colVal=specialHeaderFormat(val,srcInfo['headerformatspec'])
        else:
          colVal=val
      if (i < (len(rows)-1)) or (not (str(colVal) == 'nan') and not ('Unnamed:' in str(colVal))):
        attrs[j].append(colVal)

  drrows=[i-1 for i in rows]
  if skipRows is not None:
    skipRows = [skipRows[i]-1 for i in range(len(skipRows))]
    drrows =drrows+skipRows

  if maxRow>-1:
    drrows=drrows+[i for i in range(maxRow,len(list(df.index)))]
  if -1 in drrows:
    drrows.remove(-1)
  df.drop(df.index[drrows], inplace=True)

  if filterRows is not None:
    df = df[(df[filterRows]==coll)]

  headers=[]
  if 'specAttr' in srcInfo:
    format=srcInfo['specAttr']
    newattrs=specialAttrFormat(attrs, format)
    headers = formatForBQ(newattrs, lc=True)
  else:
    headers = formatForBQ(attrs,lc=True)
  df.columns=headers
  df.index=list(df.iloc[:,patientIdRow])

  headerSet = {}
  for i in range(len(headers)):
    headerSet[headers[i]]={"attrs":attrs[i],"colNo":i}

  if ("reindex" in srcInfo) and ("needed" in srcInfo["reindex"]) and srcInfo["reindex"]["needed"]:
    uniques = srcInfo["reindex"]["uniques"]
    df_new=pd.DataFrame()
    #df_new.index=df.index
    #df_new.columns=list(df.columns)
    newInd={}
    pos=0
    for i in range(df.shape[0]):
      curInd=df.index[i]
      if not (curInd in newInd):
        df_new = df_new.append(df.iloc[i])
        newInd[curInd]=pos
        pos=pos+1
      else:
        for colInd in range(len(df.columns)):
          if not (colInd==patientIdRow) and not (colInd in uniques):
            curVal= df_new.iloc[newInd[curInd]][colInd]
            addVal =  df.iloc[i][colInd]
            #df_new.loc[curInd, list(df.columns)[colInd]]=10
            df_new.loc[curInd, list(df.columns)[colInd]]=str(curVal)+", "+str(addVal)
    df = pd.concat([df_new])
  try:
    df[df.columns[patientIdRow]] = df[df.columns[patientIdRow]].astype('Int64')
  except:
    df[df.columns[patientIdRow]] = df[df.columns[patientIdRow]].astype('str')
  return [headerSet,df,sheetnm]

def specialHeaderFormat(val,format):
  nval=''
  if format=="lidc":
    nval= " ".join([x for x in val.split('\n') if not ("=" in x)])
  else:
    nval=val
  return nval


def specialAttrFormat(attrs, format):
  newatts=[]
  if format == 'duke':
    for attr_list in attrs:
      newset=[attr_list[0],attr_list[1]]
      optCol=attr_list[2]
      if '{' in optCol:
        val=optCol.split('{')[0]
      elif ' (' in optCol:
        val=optCol.split(' (')[0]
      elif optCol.startswith('('):
        val=optCol.split('(')[0]
      elif '=' in optCol:
        val=''
      else:
        val=optCol
      newset.append(val)
      newatts.append(newset)
  return newatts


def formatForBQ(attrs, lc=False):
  patt=re.compile(r"[a-zA-Z_0-9]")
  justNum=re.compile(r"[0-9]")
  headcols=[]
  for i in range(len(attrs)):
    headSet=[attrs[i][j] for j in range(len(attrs[i])) if len(attrs[i][j])>0]
    header='_'.join(str(k) for k in headSet)
    header=header.replace('/','_')
    header=header.replace('-', '_')
    header=header.replace(' ', '_')
    normHeader = ''
    for i in range(len(header)):
      if bool(patt.search(header[i])):
        normHeader = normHeader + header[i]
    if (len(normHeader) > 0) and bool(justNum.search(normHeader[0])):
      normHeader='c_'+normHeader
    if lc:
      normHeader = normHeader.lower()


    headcols.append(normHeader)
  return headcols

def mergeAcrossAttr(clinJson, coll):
  mbatch = clinJson[coll]['mergeBatch']
  headers= {}
  ptIdSeq=[]
  ptId = mbatch[0]['ptId'][0][1]
  try:
    new_df=pd.concat([mbatch[0]['df']])
  except:
    print("could not concate! "+coll)
    new_df = mbatch[0]['df']
  for i in range(len(mbatch)):
    ptIdSeq.append(mbatch[i]['ptId'])
    cptRow = mbatch[i]['ptId'][0][0]
    cptId = mbatch[i]['ptId'][0][1]
    cheaders = mbatch[i]['headers']
    for chead in cheaders:
      if not chead in headers and ((not chead == cptId) or (i == 0)):
        headers[chead] = {}
        headers[chead]['srcs']=[]
      if (chead == cptId):
        headers[ptId]['srcs'].append(mbatch[i]['headers'][chead])
      else:
        headers[chead]['srcs'].append(mbatch[i]['headers'][chead])
    if (i > 0):
      nxt_df = mbatch[i]['df']
      cols = list(nxt_df.columns)
      cols[cptRow] = ptId
      nxt_df.columns=cols
      #new_df = pd.concat([new_df, nxt_df])
      try:
        new_df = pd.concat([new_df, nxt_df])
      except:
        print("could not concate! " + coll)

  clinJson[coll]['headers'] = headers
  clinJson[coll]['ptIdSeq'] = ptIdSeq
  clinJson[coll]['df'] = new_df

def mergeAcrossBatch(clinJson,coll,ptRowIds,attrSetInd,colsAdded,csrc,offset):
  if 'mergeBatch' not in clinJson[coll]:
    clinJson[coll]['mergeBatch'] = []
  clinJson[coll]['mergeBatch'].append({})
  cList = list(clinJson[coll]['cols'][attrSetInd][0]['df'].columns)
  ptRow = cList[ptRowIds[0]+colsAdded]
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['ptId'] = []
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['ptId'].append([ptRowIds[0],ptRow])

  clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'] = {}

  for header in clinJson[coll]['cols'][attrSetInd][0]['headers']:
    clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header] = [clinJson[coll]['cols'][attrSetInd][0]['headers'][header]]
    clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header][0]['filenm'] = clinJson[coll][csrc][attrSetInd][0]['filenm']
    if 'sheet' in clinJson[coll][csrc][attrSetInd][0]:
      clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header][0]['sheet'] = clinJson[coll][csrc][attrSetInd][0]['sheet']
    if 'sheetnm' in clinJson[coll][csrc][attrSetInd][0]:
      clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header][0]['sheetnm'] = clinJson[coll][csrc][attrSetInd][0]['sheetnm']
    clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header][0]['batch'] = 0

  clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs']=[]
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs'].append([])
  if 'archive' in clinJson[coll][csrc][attrSetInd][0]:
    clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs'][0]=clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs'][0]+clinJson[coll][csrc][attrSetInd][0]['archive']
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs'][0].append(clinJson[coll][csrc][attrSetInd][0]['filenm'])

  df_all_rows =  pd.concat([clinJson[coll]['cols'][attrSetInd][0]['df']])


  for batchSetInd in range(1,len(clinJson[coll]['cols'][attrSetInd])):
    nList = list(clinJson[coll]['cols'][attrSetInd][batchSetInd]['df'].columns)
    cptRow = nList[ptRowIds[batchSetInd]+colsAdded]
    clinJson[coll]['mergeBatch'][attrSetInd]['ptId'].append([ptRowIds[batchSetInd],cptRow])
    if not ptRow == cptRow:
      print("Different patientColumn! "+coll)
    for colInd in range(len(nList)):
      col = nList[colInd]
      if not (col in cList) and not (col == cptRow):
        cList.append(col)
    for header in clinJson[coll]['cols'][attrSetInd][batchSetInd]['headers']:
      if not header in clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'] and not header == cptRow:
        clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header] = []
      ckey = header
      if header == cptRow:
        ckey = ptRow
        clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][ptRow].append(clinJson[coll]['cols'][attrSetInd][batchSetInd]['headers'][cptRow])
      else:
        clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][header].append(clinJson[coll]['cols'][attrSetInd][batchSetInd]['headers'][header])
      curInd=len(clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][ckey])-1
      clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][ckey][curInd]['filenm']=clinJson[coll]['srcs'][attrSetInd][batchSetInd]['filenm']

      if 'sheet' in clinJson[coll][csrc][attrSetInd][batchSetInd]:
        clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][ckey][curInd]['sheet'] = clinJson[coll]['srcs'][attrSetInd][batchSetInd]['sheet']
      if 'sheetnm' in clinJson[coll]['srcs'][attrSetInd][batchSetInd]:
        clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][ckey][curInd]['sheetnm'] = clinJson[coll]['srcs'][attrSetInd][batchSetInd]['sheetnm']
      clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][ckey][curInd]['batch']=batchSetInd


    clinJson[coll]['mergeBatch'][attrSetInd+offset][csrc].append([])
    if 'archive' in clinJson[coll][csrc][attrSetInd][batchSetInd]:
      clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs'][batchSetInd] = clinJson[coll]['mergeBatch'][attrSetInd+offset]['srcs'][batchSetInd] + clinJson[coll]['srcs'][attrSetInd][batchSetInd]['archive']
    clinJson[coll]['mergeBatch'][attrSetInd+offset][csrc][batchSetInd].append(clinJson[coll][csrc][attrSetInd][batchSetInd]['filenm'])
    # join data frames
    new_df = clinJson[coll]['cols'][attrSetInd][batchSetInd]['df']

    #make sure joining df is using the same patientId column name as the original
    colList=list(new_df.columns)
    colList[ptRowIds[batchSetInd]+colsAdded] = ptRow
    new_df.columns = colList
    df_all_rows=pd.concat([df_all_rows,new_df])
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['cList'] = cList
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['df'] = df_all_rows
  #clinJson[coll]['mergeBatch'][attrSetInd]['headers']['source_batch']=[]
  clinJson[coll]['mergeBatch'][attrSetInd+offset]['headers'][SOURCE_BATCH_COL]=[{'attrs':[SOURCE_BATCH_LABEL]}]

def export_meta_to_json(clinJson,filenm_meta,filenm_summary,collecs):

  hist ={}
  table_id = DEFAULT_PROJECT + "." + LAST_DATASET + '.table_metadata'
  if not LAST_DATASET == DEFAULT_DATASET:
    getHist(hist, table_id)
  metaArr=[]
  sumArr=[]
  for coll in collecs:
    colldir=coll.replace('/','_').replace(':','_')
    if 'dataset' in clinJson[coll]:
      dataset=clinJson[coll]['dataset']
    else:
      dataset = DEFAULT_DATASET

    if 'project' in clinJson[coll]:
      project = clinJson[coll]
    else:
      project = DEFAULT_PROJECT

    if ('mergeBatch' in clinJson[coll]):
      for k in range(len(clinJson[coll]['mergeBatch'])):
        if 'ptId' in clinJson[coll]['mergeBatch'][k]:
          sumDic = {}
          curDic=clinJson[coll]['mergeBatch'][k]
          curDf=clinJson[coll]['mergeBatch'][k]['df']
          dtypeL=list(curDf.dtypes)
          ptId=curDic['ptId'][0][0]
          ptCol=curDic['ptId'][0][1]

          if ('tabletypes' in clinJson[coll]) and (k<len(clinJson[coll]['tabletypes'])):
            suffix=list(clinJson[coll]['tabletypes'][k].keys())[0]
            table_description = clinJson[coll]['tabletypes'][k][suffix]
          elif ('tabletypes2' in clinJson[coll]):
            if 'tabletypes' in clinJson[coll]:
              offset = len(clinJson[coll]['tabletypes'])
            else:
              offset = 0
            suffix = list(clinJson[coll]['tabletypes2'][k-offset].keys())[0]
            table_description = clinJson[coll]['tabletypes2'][k-offset][suffix]

          else:
            suffix=DEFAULT_SUFFIX
            table_description=DEFAULT_DESCRIPTION
          collection_id = coll
          table_name = collection_id + '_' + suffix
          full_table_name=DATASET_PATH+'.'+table_name
          try:
            post_process_src = './'+DESTINATION_FOLDER+'/'+clinJson[coll]['mergeBatch'][k]['outfile']
            post_process_src_sp = post_process_src.split('/')
            post_process_src_tr = post_process_src_sp[len(post_process_src_sp)-1]
            post_process_src_current_md5 = get_md5(post_process_src)

          except:
            pass
          num_batches = len(clinJson[coll]['mergeBatch'][k]['srcs'])

          src_info = []
          for src in clinJson[coll]['mergeBatch'][k]['srcs']:
            nsrc = {}
            nsrc['srcs'] = src
            rootfile = ORIGINAL_SRCS_PATH + colldir + '/' + src[0]
            nsrc['md5'] = get_md5(rootfile)
            src_info.append(nsrc)

          sumDic['collection_id'] = collection_id
          sumDic['table_name'] = full_table_name
          sumDic['table_description'] = table_description
          sumDic['post_process_src'] = post_process_src_tr
          sumDic['number_batches'] = num_batches

          if table_name in hist:
            for nkey in hist[table_name]:
              if (nkey not in sumDic) and not (nkey == 'source_info'):
                sumDic[nkey] = hist[table_name][nkey]
            hist_post_process_src_sp = hist[table_name]['post_process_src'].split('/')
            hist_post_process_src_tr = hist_post_process_src_sp[len(hist_post_process_src_sp) - 1]
            if (hist_post_process_src_tr != post_process_src_tr) or (post_process_src_current_md5 != hist[table_name]['post_process_src_updated_md5']):
              sumDic['idc_version_table_prior'] = hist[table_name]['idc_version_table_updated'].rsplit('/')[0]
              sumDic['post_process_src_prior_md5'] = hist[table_name]['post_process_src_updated_md5'].rsplit('/')[0]

              sumDic['idc_version_table_updated'] = CURRENT_VERSION
              sumDic['post_process_src_updated_md5'] = post_process_src_current_md5
              sumDic['table_updated_datetime'] = str(datetime.now(pytz.utc))


              '''for i in range(len(src_info)):
                if (i < len(hist[table_name]['source_info'])) and (src_info[i]['srcs'][0] == hist[table_name]['source_info'][i]['srcs'][0]):
                  src_info[i]['added_md5'] = hist[table_name]['source_info'][i]['added_md5']
                  if src_info[i]['update_md5'] == hist[table_name]['source_info'][i]['update_md5']:
                    src_info[i]['prior_md5'] = hist[table_name]['source_info'][i]['prior_md5']
                  else:
                    src_info[i]['prior_md5'] = hist[table_name]['source_info'][i]['update_md5']
                else:
                  src_info[i]['added_md5'] = src_info[i]['update_md5']
                  src_info[i]['prior_md5'] = src_info[i]['prior_md5']'''
          else:
            sumDic['idc_version_table_added']=CURRENT_VERSION
            sumDic['table_added_datetime']=str(datetime.now(pytz.utc))
            #sumDic['post_process_src']=post_process_src
            sumDic['post_process_src_added_md5']=post_process_src_current_md5
            sumDic['idc_version_table_prior']=CURRENT_VERSION
            sumDic['post_process_src_prior_md5']=post_process_src_current_md5
            sumDic['idc_version_table_updated'] = CURRENT_VERSION
            sumDic['table_updated_datetime'] = str(datetime.now(pytz.utc))
            sumDic['post_process_src_updated_md5'] = post_process_src_current_md5
            sumDic['number_batches']=num_batches
            '''for i in range(len(src_info)):
              src_info[i]['added_md5'] = src_info[i]['update_md5']
              src_info[i]['prior_md5'] = src_info[i]['update_md5']'''

          sumDic['source_info']=src_info

          sumArr.append(sumDic)
          for i in range(len(curDf.columns)):
            ndic = {}
            if (str(curDf.columns[i]) == str(ptCol)):
              ndic['case_col']='yes'
            else:
              ndic['case_col'] = 'no'
            ndic['collection_id'] = collection_id
            ndic['table_name'] = full_table_name
            header = curDf.columns[i]
            try:
              if (len(curDic['headers'][header])>0):
                headerD = curDic['headers'][header][0]
              else:
                headerD={}
            except:
              rrr=1
            dftype=str(dtypeL[i].name)
            try:
              ndic['column_label']=headerD['attrs'][len(headerD['attrs'])-1]
            except:
              pass
            if (dftype=='Object') or (dftype=='object'):
              dftype = 'String'

            ndic['column']=str(header)
            ndic['data_type']=dftype
            if 'dictinfo' in headerD:
              if 'column_label' in headerD['dictinfo']:
                ndic['column_label']=headerD['dictinfo']['column_label']
              if 'data_type' in headerD['dictinfo']:
                ndic['data_type'] = headerD['dictinfo']['data_type']

            if ('dictinfo' in headerD) and ('values' in headerD['dictinfo']):
              ndic['values'] = headerD['dictinfo']['values']
              ndic['values_source'] ='provided dictionary'
              for val in ndic['values']:
                if val['option_code'].lower() == 'nan':
                  val['option_code'] = '\"' + val['option_code'] + '\"'
            elif 'uniques' in headerD:
              num_values=len(headerD['uniques'])
              if (num_values<21):
                ndic['values_source'] = 'derived from inspection of values'
                #ndic['uniques'] = headerD['uniques']
                ndic['values']=[]
                for val in headerD['uniques']:
                  cval=str(val)
                  if cval.lower()=='nan':
                    cval='\"'+cval+'\"'
                  ndic['values'].append({'option_code':cval})
              headerD.pop('uniques')
            ndic['original_column_headers'] = []
            ndic['files'] = []
            ndic['column_numbers'] = []
            ndic['sheet_names'] = []
            ndic['batch'] = []
            #sheetnms=[]
            for headerInfo in curDic['headers'][header]:
              ndic['original_column_headers'].append( str(headerInfo['attrs']))
              if not (header == SOURCE_BATCH_COL) and not (header == DICOM_COL):
                ndic['column_numbers'].append(headerInfo['colNo'])
                ndic['batch'].append(headerInfo['batch'])
                if 'sheetnm' in headerInfo:
                  ndic['sheet_names'].append(headerInfo['sheetnm'])
                ndic['files'].append( headerInfo['filenm'])
            metaArr.append(ndic)
  f=open(filenm_summary,'w')
  json.dump(sumArr,f)
  f.close()
  f = open(filenm_meta, 'w')
  json.dump(metaArr, f)
  f.close()


def reform_case(case_id, colec,type):
  if type == "same":
    ret = case_id
  elif type == "breast":
    ret = case_id.replace("BreastDX","BreastDx")
  elif type == "acrin 6698":
    ret = "ACRIN-6698-"+case_id
  elif type == "acrin format":
    ret=colec+'-'+case_id.rjust(3,'0')
  elif type == "acrin flt format":
    ret=colec+'_'+case_id.rjust(3,'0')
  elif type == 'cc':
    prts=case_id.split('_')
    prts2=prts[1].split('-')
    ret=prts[0]+'-'+prts2[0]+prts2[1].rjust(2,'0')
  elif type == "switch dash":
    ret=case_id.replace('_','-')
  elif type == "3DCT-RT":
    ret="HN_P"+case_id.rjust(3,'0')
  elif type=="ispy":
    ret="ISPY1_"+case_id
  elif type=="ispy2":
    ret="ISPY2-"+case_id
  elif type=="lung_pt":
    ret = "Lung_Dx-"+case_id
  elif type=='add colec':
    ret=colec+'-'+case_id
  elif type=='remind':
    ret='ReMIND-'+case_id.zfill(3)
  elif type=='ea1141':
    ret='EA1141-'+case_id
  return ret

def add_tcia_case_id(mergeB, tcia_coll,type):
  colId=mergeB['ptId'][0][1]
  df=mergeB['df']
  ncaseA=df[colId].apply(lambda x: reform_case(str(x),tcia_coll,type))
  df.insert(0,'dicom_patient_id',ncaseA)
  mergeB['headers']['dicom_patient_id']=[{'attrs':[DICOM_LABEL]}]


def parse_acrin_collection(clinJson,coll):
  webapp_coll=clinJson[coll]['idc_webapp']
  clinJson[coll]['mergeBatch']=[]
  clinJson[coll]['tabletypes']=[]
  colldir=coll.replace('/','_')
  colldir=colldir.replace(':','_')
  curDir= ORIGINAL_SRCS_PATH  + colldir


  if 'uzip' in clinJson[coll]:
    if 'udir' in clinJson[coll]:
      [shutil.rmtree(curDir+ '/'+d, ignore_errors=True) for d in clinJson[coll]['udir']]
    for zpfile in clinJson[coll]['uzip']:
      zpfile = curDir + '/' + zpfile
      with zipfile.ZipFile(zpfile) as zip_ref:
        kk=1
        zip_ref.extractall(ORIGINAL_SRCS_PATH  + colldir)

    internalDirs=[d for d in os.listdir(curDir) if os.path.isdir(os.path.join(curDir,d))]
    [os.rename(curDir+ '/'+d, curDir+ '/'+d.replace('/','_').replace(':','_')) for d in internalDirs]

    if 'udir' in clinJson[coll]:
      cdir = clinJson[coll]['udir'][0]
    else:
      cdir= path.splitext(clinJson[coll]['uzip'][0])[0]
    npath = curDir + '/' +cdir
    ofiles = [f for f in listdir(npath) if path.isfile(path.join(npath,f))]
    dictFile = [f for f in ofiles if 'Dictionary' in f][0]

    formFileA = [f for f in ofiles if ('Form' in f) and ('.xls' in f)]
    if len(formFileA)>0:
      formFile=npath+'/'+formFileA[0]
    else:
      formFile = None
    dictFile= npath+'/'+dictFile
    if 'dictfile' in clinJson[coll]:
      dictFile=curDir + '/' + clinJson[coll]['dictfile']
    parser=clinical.acrin_forms.DictionaryReader(dictFile,formFile)
    parser.parse_dictionaries()
    dict_names = parser.get_dictionary_names()
    for form_id in dict_names:
      desc = parser.get_dictionary_desc(form_id)
      cdict = parser.get_dictionary(form_id)
      reformCdict={}
      for celem in cdict:
        celem['variable_name']=celem['variable_name'].lower()
        celem['column'] = celem['variable_name']
        del celem['variable_name']
        celem['column_label'] = celem['variable_label']
        del celem['variable_label']
        reformCdict[celem['column']]=celem



      srcf=npath+'/'+form_id+'.csv'
      if path.exists(srcf):
        #print(srcf)

        clinJson[coll]['tabletypes'].append({form_id:desc})
        df = pd.read_csv(srcf)
        ptId = [[0, df.columns[0].lower()]]
        df.insert(0, 'source_batch', 0)
        #headers['source_batch'] = {'attrs': ['NA'], 'colNo': -1}

        colnames = [list(df.columns)]

        if len(clinJson[coll]['uzip'])>1:
          if 'udir' in clinJson[coll]:
            ccdir = clinJson[coll]['udir'][1]
          else:
            ccdir = path.splitext(clinJson[coll]['uzip'][1])[0]
          osrcf= ORIGINAL_SRCS_PATH + colldir + '/' + ccdir + '/' +form_id+'.csv'
          df2 = pd.read_csv(osrcf)
          ptId.append([0, df2.columns[0].lower()])
          df2.insert(0, 'source_batch', 1)
          #headers['source_batch'] = {'attrs': ['NA'], 'colNo': -1}
          colnames.append(list(df.columns))
          df = pd.concat([df, df2])
        #shutil.copy2(srcf,destf)
        recastDataFrameTypes(df, 0)
        destf = DESTINATION_FOLDER +'/' + webapp_coll + '_' + form_id + '.csv'
        #df.to_csv(destf, index=False)
        ndic={}
        ndic['df']=df
        ndic['ptId']=ptId
        ndic['headers'] ={}
        orig_names=list(df.columns)
        norm_names=[df.columns[k].lower() for k in range(len(df.columns))]
        df.columns=norm_names
        for k in range(len(df.columns)):
          headval = df.columns[k]
          ndic['headers'][headval] = []
          #df.columns[k]=df.columns[k].lower()
          orig_nm=orig_names[k]
          for kk in range(len(colnames)):
            if orig_nm in colnames[kk]:
              ind = colnames[kk].index(orig_nm)
              hndic={}
              if orig_nm == SOURCE_BATCH_COL:
                hndic['attrs']=[SOURCE_BATCH_LABEL]
              else:
                hndic['attrs']=[orig_nm]
                hndic['colNo'] = ind
                hndic['sheet'] = 0
              if kk == 0:
                hndic['filenm'] = cdir+'/'+form_id + '.csv'
              else:
                hndic['filenm'] = ccdir + '/' + form_id + '.csv'
              hndic['batch'] =kk
            if (headval in reformCdict) and (len(ndic['headers'][headval])==0):
              hndic['dictinfo'] = reformCdict[headval]
            ndic['headers'][headval].append(hndic)
        ndic['srcs'] = [[clinJson[coll]['uzip'][0], cdir + '/' + form_id + '.csv']]
        if (len(clinJson[coll]['uzip'])>1):
          ndic['srcs'].append([clinJson[coll]['uzip'][1],ccdir + '/' + form_id + '.csv'])
        ndic['outfile']=webapp_coll + '_' + form_id + '.csv'
        if 'tcia_api' in clinJson[coll]:
          add_tcia_case_id(ndic, clinJson[coll]['tcia_api'], clinJson[coll]['case_id'])
        else:
          add_tcia_case_id(ndic, None, clinJson[coll]['case_id'])
        ndic['df'].to_csv(destf, index=False)
        ndic['source_batch']=[]
        clinJson[coll]['mergeBatch'].append(ndic)

  pass


def parse_conventional_collection(clinJson,coll,csrc,tbltypes):
  #clinJson[coll]['idc_webapp']=coll
  colldir = coll.replace('/','_').replace(':','_')
  if ('uzip' in clinJson[coll]) and (csrc == 'srcs'):
    zpfile = ORIGINAL_SRCS_PATH + colldir + '/' + clinJson[coll]['uzip']
    with zipfile.ZipFile(zpfile) as zip_ref:
      zip_ref.extractall(ORIGINAL_SRCS_PATH + colldir)
  if csrc in clinJson[coll]:
    if 'mergeBatch' in clinJson[coll]:
      offset = len(clinJson[coll]['mergeBatch'])
    else:
      offset = 0
    if 'cols' in clinJson[coll]:
      offset2 = len(clinJson[coll]['cols'])
    else:
      clinJson[coll]['cols'] = []
      offset2 = 0

    for attrSetInd in range(len(clinJson[coll][csrc])):
      ptRowIds = []
      #cohortSeries = clinJson[coll]['srcs'][attrSetInd]
      clinJson[coll]['cols'].append([])
      wJson = False
      for batchSetInd in range(len(clinJson[coll][csrc][attrSetInd])):
        clinJson[coll]['cols'][attrSetInd+offset2].append([])
        clinJson[coll]['cols'][attrSetInd+offset2][batchSetInd] = {}
        srcInfo = clinJson[coll][csrc][attrSetInd][batchSetInd]


        #print("strcInfo " + str(srcInfo))
        if not ('type' in srcInfo) or not (srcInfo['type'] == 'json'):
          [headers, df, sheetnm] = processSrc(ORIGINAL_SRCS_PATH, colldir, srcInfo, coll)
          #df['source_batch'] = batchSetInd
          df.insert(0, 'source_batch', batchSetInd)
          headers['source_batch'] = {'attrs':['NA'], 'colNo':-1}
          # attrs.append([attr])
          srcInfo['sheetnm']=sheetnm
          clinJson[coll]['cols'][attrSetInd+offset2][batchSetInd]['headers'] = headers
          clinJson[coll]['cols'][attrSetInd+offset2][batchSetInd]['df'] = df
        else:
          wJson = True

        patientIdRow=0
        if 'patientIdRow' in srcInfo:
          patientIdRow=srcInfo['patientIdRow']
        elif 'patientIdCol' in srcInfo:
          patientIdRow=headers[srcInfo['patientIdCol']]['colNo']
        ptRowIds.append(patientIdRow)

      if not wJson in clinJson[coll]:
        colsAdded=1
        mergeAcrossBatch(clinJson, coll, ptRowIds, attrSetInd, colsAdded,csrc,offset)
        recastDataFrameTypes(clinJson[coll]['mergeBatch'][attrSetInd+offset]['df'],
                             clinJson[coll]['mergeBatch'][attrSetInd+offset]['ptId'][0][0])
        analyzeDataFrame(clinJson[coll]['mergeBatch'][attrSetInd+offset])
        suffix = DEFAULT_SUFFIX
        if tbltypes in clinJson[coll]:
          suffix = list(clinJson[coll][tbltypes][attrSetInd].keys())[0]
        #nm = clinJson[coll]['idc_webapp'] + '_' + suffix
        nm = coll + '_' + suffix
        clinJson[coll]['mergeBatch'][attrSetInd+offset]['outfile'] = nm + '.json'
        if 'tcia_api' in clinJson[coll]:
          if 'case_id' in srcInfo:
            add_tcia_case_id(clinJson[coll]['mergeBatch'][attrSetInd + offset], clinJson[coll]['tcia_api'],srcInfo['case_id'])
          else:
            add_tcia_case_id(clinJson[coll]['mergeBatch'][attrSetInd+offset], clinJson[coll]['tcia_api'], clinJson[coll]['case_id'])
        else:
          if 'case_id' in srcInfo:
            add_tcia_case_id(clinJson[coll]['mergeBatch'][attrSetInd + offset], None, srcInfo['case_id'])
          else:
            add_tcia_case_id(clinJson[coll]['mergeBatch'][attrSetInd+offset], None, clinJson[coll]['case_id'])


        write_dataframe_to_json(DESTINATION_FOLDER, nm, clinJson[coll]['mergeBatch'][attrSetInd+offset]['df'])


def nlst_handler(filenm, sheetNo, data_dict):
  wb = openpyxl.load_workbook(filename=filenm)
  ws = wb.worksheets[sheetNo]
  mxRow = ws.max_row
  cellBnds = []
  for rng in ws.merged_cells.ranges:
    if rng.max_col == 1:
      cellBnds.append([rng.min_row, rng.max_row])

  cellBnds = sorted(cellBnds, key=lambda x: x[0])
  nBnds = []
  mn = cellBnds[0][0]
  for i in range(2, mn):
    nBnds.append([i, i])
  for i in range(0, len(cellBnds) - 1):
    nBnds.append(cellBnds[i])
    mn = cellBnds[i][1]
    mx = cellBnds[i + 1][0]
    for j in range(mn + 1, mx):
      nBnds.append([j, j])
  nBnds.append(cellBnds[len(cellBnds) - 1])
  mx = cellBnds[len(cellBnds) - 1][1]

  for i in range(mx + 1, mxRow + 1):
    nBnds.append([i, i])
  for bnd in nBnds:
    # for row_ind in range(bnd[0],bnd[1]):
    column = "".join([str(ws.cell(row=x, column=1).value) for x in range(bnd[0], bnd[1] + 1) if
                      ws.cell(row=x, column=1).value is not None])
    col_orig=column
    column = formatForBQ([[column]], True)[0]
    column_label = "".join([str(ws.cell(row=x, column=2).value) for x in range(bnd[0], bnd[1] + 1) if
                            ws.cell(row=x, column=2).value is not None])
    column_label_add = "".join([str(ws.cell(row=x, column=3).value) for x in range(bnd[0], bnd[1] + 1) if
                                ws.cell(row=x, column=3).value is not None])
    if len(column_label_add) > 0:
      column_label = column_label + ': ' + column_label_add

    cols=[]
    mtch=re.search('[0-9]-[0-9]$',col_orig)
    if mtch is None:
      cols=[column]
    else:
      mtchStr=mtch.group(0)
      croot=column[:len(column)-3]
      strt=int(mtchStr[0])
      end=int(mtchStr[2])
      for j in range(strt,end+1):
        cols.append(croot+str(j))

    for column in cols:
      data_dict[column] = {}
      data_dict[column]['label'] = column_label

      for row_ind in range(bnd[0], bnd[1] + 1):
        if (ws.cell(row=row_ind, column=4).value is not None) and ('=' in ws.cell(row=row_ind, column=4).value):
          vals = ws.cell(row=row_ind, column=4).value.split('=')
          option_code = vals[0].strip('"').strip("'")
          option_description = vals[1].strip('"').strip("'")
          if not ('opts' in data_dict[column]):
            data_dict[column]['opts'] = []
          data_dict[column]['opts'].append({"option_code": option_code, "option_description": option_description})
  return

def parse_dict(fpath,collec,ndic,indx,coll):
  data_dict={}
  colldir = coll.replace('/', '_').replace(':', '_')
  filenm=fpath + colldir + '/' +ndic["filenm"]
  sheetNo=0
  if "sheet" in ndic:
    sheetNo=ndic["sheet"]
  skipRows=None
  if "skipRows" in ndic:
    skipRows= ndic["skipRows"]
  header = 0
  if "header" in ndic:
    if ndic["header"]=="None":
      header = None
    else:
      header = ndic["header"]

  extension = path.splitext(filenm)[1]
  engine = 'xlrd'
  if extension == '.xlsx':
    engine = 'openpyxl'
  elif extension == '.xlsb':
    engine = 'pyxlsb'
  df = []
  dc =[]
  if extension == '.csv':
    df = pd.read_csv(filenm, header=header, skiprows=skipRows,keep_default_na=False)
    sheetnm = ''
  elif extension == '.docx':

    dc = [docx2python(filenm).document[0][0][0]]
    if ("filenm2" in ndic):
      filenm2 = fpath + colldir + '/' + ndic["filenm2"]
      dc.append(docx2python(filenm2).document[0][0][0])

  else:
    dfi = pd.read_excel(filenm, engine=engine, sheet_name=None, skiprows=skipRows, header=header,keep_default_na=False)
    if not isinstance(sheetNo, list):
      sheetnm = list(dfi.keys())[sheetNo]
      df = dfi[sheetnm]
    rr=1
  if (ndic["form"] =="adrenal"):
      for index, row in df.iterrows():
        column=row['Column Name']
        column=formatForBQ([[column]],True)[0]
        label=row['Description']
        opts=row['Potential Values'].split(',')

        data_dict[column] = {}
        data_dict[column]['label'] = label
        if len(opts)>0:
          data_dict[column]['opts'] = []
          for opt in opts:
            nopt=opt.strip()
            data_dict[column]['opts'].append({"option_code": nopt})

  elif (ndic["form"] =="colorectal"):
    for index, row in df.iterrows():
      column = row['Variable Name']
      column = formatForBQ([[column]], True)[0]
      label = row['Description']
      opts = re.split(r"[|,]\s*(?![^()]*\))", row['Values'])
      #opts = row['Values'].split('|')

      data_dict[column] = {}
      data_dict[column]['label'] = label
      if len(opts) > 0:
        data_dict[column]['opts'] = []
        for opt in opts:
          nopt = opt.strip()
          optA=nopt.split(' - ')
          optA=[opt.strip() for opt in optA]
          if len(optA)>1:
            data_dict[column]['opts'].append({"option_code": optA[0], "option_description":optA[1]})
          else:
            data_dict[column]['opts'].append({"option_code": optA[0]})

  elif (ndic["form"]=="duke"):
    headers =collec['mergeBatch'][0]['headers']
    for head in headers:
      if ('attrs' in headers[head][0]) and (len(headers[head][0]['attrs']) == 3):
        column=head
        data_dict[column]={}
        optCol = headers[head][0]['attrs'][2]
        if optCol.startswith('('):
          data_dict[column]['label']=optCol
        else:
          data_dict[column]['label'] =''
        if '{' in optCol:
          optCol = optCol.split('{')[1]
          optCol=optCol.replace('}','')

        if '=' in optCol:
          optColA=optCol.split('=')

          if (len(optColA)>1):
            data_dict[column]['opts'] = []
            optCode=[optColA[0]]
            optDesc=[]
            for k in range(1,len(optColA)-1):
              nxtset=re.split(r"[;,\n]+\s*",optColA[k],1)
              if len(nxtset)==2:
                optDesc.append(nxtset[0])
                optCode.append(nxtset[1])
              else:
                nxtset = re.split(r"\s+", optColA[k], 1)
                optDesc.append(nxtset[0])
                optCode.append(nxtset[1])
            optDesc.append(optColA[len(optColA)-1])
            #possible optCode and optDesc are switched. code showed be 'shorter' than desc
            optDescL="".join(optDesc)
            optCodeL = "".join(optCode)
            if (len(optCodeL)>len(optDescL)):
              tmp=[x for x in optDesc]
              optDesc=[x for x in optCode]
              optCode=[x for x in tmp]


            for k in range(len(optCode)):
              data_dict[column]['opts'].append({"option_code": optCode[k], "option_description": optDesc[k]})
        if (len(data_dict[column].keys())==0):
          del data_dict[column]


  elif (ndic["form"] =="ispy"):
    data_dict=parseIspyDic(df)
  elif (ndic["form"]=="ispy2"):
    for index,row in df.iterrows():
      column = ' '.join(row['FIELD'].split())
      column = formatForBQ([[column]],True)[0]
      descriptionA = row['DESCRIPTION'].split('\n')
      label = ' '.join(descriptionA[0].split())
      data_dict[column] = {}
      data_dict[column]['label']=label
      if (len(descriptionA)>1):
        data_dict[column]['opts']=[]
      for k in range(1,len(descriptionA)):
        optA = ' '.join(descriptionA[k].split()).split(':')
        if len(optA)>1:
          data_dict[column]['opts'].append({"option_code":optA[0], "option_description":optA[1]})
        else:
          data_dict[column]['opts'].append({"option_code": optA[0]})
  elif (ndic["form"]=="lidc"):
    for col in list(df.columns):
      column_label = specialHeaderFormat(col,"lidc")
      column = formatForBQ([[column_label]],True)[0]
      data_dict[column] = {}
      data_dict[column]['label'] = column_label
      opts= [x for x in col.split('\n') if '=' in x]
      if len(opts)>0:
        data_dict[column]['opts'] = []
        for op in opts:
          optA=op.split('=')
          data_dict[column]['opts'].append({"option_code": optA[0], "option_description": optA[1]})
  elif (ndic["form"]=="hcc_tace"):
    spec={"Y = 1 N = 0", "1=Male, 2=Female"}
    for index, row in df.iterrows():
      column = formatForBQ([[row[0]]], True)[0]
      column_label = row[1]
      if column_label in spec:
        column_label = row[0]+": "+row[1]
      data_dict[column] = {}
      data_dict[column]['label'] = column_label
  elif (ndic["form"]=="covid"):
    for index, row in df.iterrows():
      column = formatForBQ([[row['column_name']]], True)[0]
      column_label =row['description']
      data_dict[column] = {}

      if len(row['column_counts'])>0:
          data_dict[column]['opts'] = []
          optJ=json.loads(row['column_counts'].replace('\'','"').replace('False','"False"').replace('True','"True"'))
          for nkey in optJ:
            data_dict[column]['opts'].append({"option_code": nkey})
      data_dict[column]['label'] = column_label

  elif (ndic["form"]=="men"):
    for index, row in df.iterrows():
      column=row['Field']
      column_label=row['Definition']
      data_dict[column] = {}
      data_dict[column]['label']=column_label

  elif (ndic["form"]=="PROSTATEx"):
    column=''
    for dci in dc:
      for nxtStr in dci:
        if re.search("^--\\t", nxtStr):
          nxtStr = nxtStr.replace("--\t", "")
          nxtA=[cstr.strip() for cstr in nxtStr.split(chr(8211))]
          column=formatForBQ([[nxtA[0]]], True)[0]
          if (len(nxtA)>1):
            data_dict[column] = {'label':nxtA[1]}
        elif re.search("^\\t--\\t", nxtStr):
          nxtStr = nxtStr.replace("\t--\t", "")
          nxtA = [cstr.strip() for cstr in re.split('-|'+chr(8211),nxtStr)]
          if not (column in data_dict):
            data_dict[column]={}
          if not ('nopts' in data_dict[column]):
            data_dict[column]['nopts'] = {}
            data_dict[column]['opts'] =[]
          if not nxtA[0] in data_dict[column]['nopts']:
            data_dict[column]['nopts'][nxtA[0]]=1
            data_dict[column]['opts'].append({"option_code": nxtA[0], "option_description": nxtA[1]})
    for col in data_dict:
      if 'nopts' in data_dict[col]:
        del data_dict[col]['nopts']

  elif (ndic["form"]=="nlst"):
    data_dict={}
    nlst_handler(filenm, sheetNo, data_dict)

  elif (ndic["form"]=="nlst2"):
    data_dict={}
    for num in sheetNo:
      nlst_handler(filenm, num, data_dict)

  elif (ndic["form"]=="ea1141"):
    colSet = set()
    headers = collec['mergeBatch'][0]['headers']
    for head in headers:
      if ('attrs' in headers[head][0]):
        col=headers[head][0]['attrs'][0]
        colSet.add(col)

    column=''
    for index, row in df.iterrows():
      if (row[0] in colSet):
        column = formatForBQ([[row[0]]], True)[0]
        description=row[1]
        data_dict[column] = {}
        data_dict[column]['label'] = description
        data_dict[column]['opts'] = []
      elif (len(row[0])>0):
        column=''
      if len(column)>0:
        if len(row[2])>0 and ("=" in row[2]):
          optA=row[2].split('\n')
          for optS in optA:
            if ("=" in optS):
              opts=optS.split("=")
              data_dict[column]['opts'].append({"option_code": opts[0], "option_description": opts[1]})
    for column in data_dict:
      if ('opts' in data_dict[column]) and (len(data_dict[column]['opts'])==0):
        del data_dict[column]['opts']

  elif (ndic["form"]=="remind"):
    for index, row in df.iterrows():
      column = formatForBQ([[row[0]]], True)[0]
      data_dict[column]={}
      data_dict[column]['label'] = row[1]


  btch = collec['mergeBatch'][indx]

  for nkey in btch['headers']:
    if nkey in data_dict:
      btch['headers'][nkey][0]['dictinfo'] = {}
      btch['headers'][nkey][0]['dictinfo']['column'] = nkey
      if 'label' in data_dict[nkey]:
        btch['headers'][nkey][0]['dictinfo']['column_label'] = data_dict[nkey]['label']
      if 'opts' in data_dict[nkey]:
        btch['headers'][nkey][0]['dictinfo']['values'] = data_dict[nkey]['opts']


def add_from_archive():
  alist = listdir(ARCHIVE_FOLDER)
  for adir in alist:
    if not (adir == '.DS_Store') and not (adir == 'bamf'):
      cdir = ARCHIVE_FOLDER+'/'+adir
      destdir = ORIGINAL_SRCS_PATH+adir
      srclist = listdir(cdir)
      dirpath = Path(destdir)
      if not dirpath.exists():
        mkdir(destdir)

      for src in srclist:
        fsrc = cdir+'/'+src
        shutil.copy(fsrc,destdir)
        if src.endswith('.zip'):
          ndest = destdir + '/' + src
          shutil.unpack_archive(ndest, destdir)

  bamfdir = ARCHIVE_FOLDER+'bamf'
  for bfile in settings.BAMF_SET:
    srcfile = bamfdir+'/'+bfile
    colecs = settings.BAMF_SET[bfile]

    for colec in colecs:
      destdir = ORIGINAL_SRCS_PATH+colec
      if not os.path.exists(destdir):
        os.mkdir(destdir)
      elif not os.path.isdir(destdir):
        os.remove(destdir)
        os.mkdir(destdir)
      shutil.copy(srcfile, destdir)


if __name__=="__main__":

  add_from_archive()

  dirpath = Path(DESTINATION_FOLDER)
  clinJson = read_clin_file(NOTES_PATH + 'clinical_notes.json')
  #clinJson = read_clin_file(NOTES_PATH + 'test_notes.json')
  collec=list(clinJson.keys())
  collec.sort()

  update=False
  if (len(sys.argv)>1):
    updateNum=sys.argv[1]
    selCol=sys.argv[2]
    update=True
    if selCol in collec:
      collec=[selCol]
    else:
      collec=[]
      exit()
  else:
    if dirpath.exists() and dirpath.is_dir():
      shutil.rmtree(dirpath, ignore_errors = True)
    mkdir(dirpath)


  for collID in range(len(collec)):
    coll=collec[collID]
    if 'spec' in clinJson[coll]:
      if (clinJson[coll]['spec'] == 'ignore') or (clinJson[coll]['spec'] == 'error'):
        pass
      elif clinJson[coll]['spec'] == 'acrin':
        parse_acrin_collection(clinJson,coll)
    
    elif ('srcs' in clinJson[coll]) and ('tcia' in clinJson[coll]) and (clinJson[coll]['tcia'] == "yes"):
      print("about to parse "+coll)
      parse_conventional_collection(clinJson, coll,'srcs','tabletypes')
      if "dict" in clinJson[coll]:
        for indx in range(len(clinJson[coll]["dict"])):
          ndic=clinJson[coll]["dict"][indx]
          if ("use" in ndic) and ndic:
            parse_dict(ORIGINAL_SRCS_PATH,clinJson[coll],ndic,indx,coll)
    


    if ('srcs2' in clinJson[coll]):
      print("about to parse extra " + coll)
      parse_conventional_collection(clinJson, coll, 'srcs2', 'tabletypes2')




  clin_meta=  CURRENT_VERSION +'_column_metadata.json'
  clin_summary = CURRENT_VERSION +'_table_metadata.json'
  if update:
    clin_meta = CURRENT_VERSION + '_' + updateNum + '_column_metadata.json'
    clin_summary = CURRENT_VERSION + '_' + updateNum + '_table_metadata.json'
  export_meta_to_json(clinJson,clin_meta,clin_summary,collec)
  i=1

